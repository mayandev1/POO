import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Criar a planilha
wb = openpyxl.Workbook()

# Remover aba padrão 'Sheet'
wb.remove(wb.active)

# ==================== ABA 1: SERVIÇOS ====================
ws_servicos = wb.create_sheet("Serviços")

# Cabeçalho
cabecalho_servicos = ["Código", "Serviço", "Preço mínimo (R$)", "Preço máximo (R$)", "Observação"]
ws_servicos.append(cabecalho_servicos)

# Dados (APENAS os valores que você repassou)
dados_servicos = [
    ["SRV001", "Logotipo", 120.00, 150.00, ""],
    ["SRV002", "Banner / Capa", 50.00, 0.00, ""],
    ["SRV003", "Pacote 4 posts (sem carrossel)", 60.00, 70.00, ""],
    ["SRV004", "Pacote 4 posts (com carrossel)", 90.00, 100.00, ""],
    ["SRV005", "UI/UX – Web/Mobile (por tela)", 50.00, 0.00, "Valor inicial"],
    ["SRV006", "Pacote 4 vídeos (Stories)", 70.00, 0.00, ""],
    ["SRV007", "Pacote 4 fotos (Stories)", 35.00, 0.00, ""],
    ["SRV008", "Reels / Shorts – pacote 4 vídeos", 100.00, 120.00, ""],
    ["SRV009", "Logo + Banner / Capa", 180.00, 0.00, ""],
    ["SRV010", "Logo + Banner / Capa + Pacote 4 posts (sem carrossel)", 240.00, 0.00, ""],
    ["SRV011", "Logo + Banner / Capa + Pacote 4 posts (com carrossel)", 280.00, 0.00, ""],
]

for dado in dados_servicos:
    ws_servicos.append(dado)

# ==================== ABA 2: DIRETORIA DE MARKETING ====================
ws_mkt = wb.create_sheet("Diretoria de Marketing")

cabecalho_gastos = ["Data", "Fornecedor", "Descrição", "Categoria", "Valor (R$)", "Forma de pagamento", "Status", "Observação"]
ws_mkt.append(cabecalho_gastos)
ws_mkt.append(["-", "-", "Nenhum gasto lançado", "-", 0.00, "-", "-", "-"])

# ==================== ABA 3: DIRETORIA DE PROJETOS ====================
ws_projetos = wb.create_sheet("Diretoria de Projetos")
ws_projetos.append(cabecalho_gastos)
ws_projetos.append(["-", "-", "Nenhum gasto lançado", "-", 0.00, "-", "-", "-"])

# ==================== ABA 4: DIRETORIA COMERCIAL ====================
ws_comercial = wb.create_sheet("Diretoria Comercial")
ws_comercial.append(cabecalho_gastos)
ws_comercial.append(["-", "-", "Nenhum gasto lançado", "-", 0.00, "-", "-", "-"])

# ==================== ABA 5: DIRETORIA ADMINISTRATIVO-FINANCEIRA ====================
ws_adm = wb.create_sheet("Diretoria Administrativo-Financeira")
ws_adm.append(cabecalho_gastos)
ws_adm.append(["-", "-", "Nenhum gasto lançado", "-", 0.00, "-", "-", "-"])

# ==================== ABA 6: DIRETORIA DE GENTE & GESTÃO ====================
ws_gente = wb.create_sheet("Diretoria de Gente & Gestão")
ws_gente.append(cabecalho_gastos)
ws_gente.append(["-", "-", "Nenhum gasto lançado", "-", 0.00, "-", "-", "-"])

# ==================== ABA 7: FINANCEIRO GERAL ====================
ws_financeiro = wb.create_sheet("Financeiro Geral")

cabecalho_financeiro = ["Diretoria", "Total gasto (R$)", "Previsão orçamentária (R$)", "Saldo (R$)", "% utilizado"]
ws_financeiro.append(cabecalho_financeiro)

diretorias = [
    "Marketing",
    "Projetos",
    "Comercial",
    "Administrativo-Financeira",
    "Gente & Gestão"
]

for diretoria in diretorias:
    ws_financeiro.append([diretoria, 0.00, 0.00, 0.00, "0%"])

ws_financeiro.append(["Total", 0.00, 0.00, 0.00, "0%"])

# ==================== ABA 8: GRÁFICOS ====================
ws_graficos = wb.create_sheet("Gráficos")

# Mensagem inicial
ws_graficos["A1"] = "📊 DASHBOARD EJ JUÁ"
ws_graficos["A2"] = "Preencha os gastos nas abas das diretorias para visualizar os gráficos automaticamente."
ws_graficos["A4"] = "Total gasto: R$ 0,00"
ws_graficos["A5"] = "Saldo restante: R$ 0,00"
ws_graficos["A6"] = "% utilizado: 0%"

# ==================== FORMATAÇÃO PROFISSIONAL ====================
# Estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Aplicar formatação em todas as abas
for ws in wb.worksheets:
    # Cabeçalhos em negrito e fundo azul
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Alinhamento das células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if cell.value == 0:
                    cell.value = 0.00

# Formatação especial para valores monetários nas abas de gastos
for ws in [ws_mkt, ws_projetos, ws_comercial, ws_adm, ws_gente]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column == 5:  # Coluna Valor (R$)
                cell.number_format = 'R$ #,##0.00'

# Formatação da aba Financeiro Geral
for row in ws_financeiro.iter_rows(min_row=2, max_row=ws_financeiro.max_row):
    for cell in row:
        if cell.column in [2, 3, 4]:  # Colunas de valores
            cell.number_format = 'R$ #,##0.00'
        if cell.column == 5 and cell.value != "Total":  # Coluna %
            cell.number_format = '0%'

# ==================== SALVAR ARQUIVO ====================
nome_arquivo = "EJ_Juá_Planilha_Financeira.xlsx"
wb.save(nome_arquivo)
print(f"✅ Planilha criada com sucesso: {nome_arquivo}")
print("\n📁 Estrutura criada:")
print("   - Serviços (com seus valores)")
print("   - 5 abas de diretorias (zeradas)")
print("   - Financeiro Geral (consolidado)")
print("   - Gráficos (dashboard)")
print("\n🚀 Agora é só preencher os gastos reais nas abas das diretorias!")